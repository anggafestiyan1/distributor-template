"""Field templates management views."""
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import IntegrityError
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, DetailView, ListView, UpdateView, View

from apps.core.mixins import StaffOrAdminMixin
from .forms import (
    HeaderFieldMappingFormSet,
    StandardMasterFieldForm,
    TemplateFieldMappingFormSet,
    TemplateForm,
    TemplateVersionForm,
)
from .models import StandardMasterField, Template, TemplateVersion


# ── Standard Master Fields ────────────────────────────────────────────────────

class FieldListView(LoginRequiredMixin, StaffOrAdminMixin, ListView):
    model = StandardMasterField
    template_name = "field_templates/field_list.html"
    context_object_name = "fields"
    ordering = ["order", "name"]


class FieldCreateView(LoginRequiredMixin, StaffOrAdminMixin, CreateView):
    model = StandardMasterField
    form_class = StandardMasterFieldForm
    template_name = "field_templates/field_form.html"

    def get_success_url(self):
        return reverse("field_templates:field_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=self.request.user,
            action=ActivityLog.ACTION_CREATE,
            description=f"Created Standard Field '{self.object.display_name}' ({self.object.name})",
            target=self.object,
            details={"name": self.object.name, "data_type": self.object.data_type},
            request=self.request,
        )
        messages.success(self.request, "Field created.")
        return response


class FieldEditView(LoginRequiredMixin, StaffOrAdminMixin, UpdateView):
    model = StandardMasterField
    form_class = StandardMasterFieldForm
    template_name = "field_templates/field_form.html"
    success_url = reverse_lazy("field_templates:field_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=self.request.user,
            action=ActivityLog.ACTION_UPDATE,
            description=f"Updated Standard Field '{self.object.display_name}'",
            target=self.object,
            details={"changed_fields": list(form.changed_data)},
            request=self.request,
        )
        messages.success(self.request, "Field updated.")
        return response


class FieldDeleteView(LoginRequiredMixin, StaffOrAdminMixin, View):
    def post(self, request, pk):
        from django.db.models import ProtectedError
        field = get_object_or_404(StandardMasterField, pk=pk)
        name = field.display_name
        code_name = field.name
        pk_val = field.pk
        try:
            field.delete()
            from apps.core.services import log_activity
            from apps.core.models import ActivityLog
            log_activity(
                user=request.user,
                action=ActivityLog.ACTION_DELETE,
                description=f"Deleted Standard Field '{name}' ({code_name})",
                details={"field_id": pk_val, "name": code_name},
                request=request,
            )
            messages.success(request, f"Field '{name}' deleted.")
        except ProtectedError as e:
            messages.error(request, str(e).split("set()")[0].strip().rstrip(","))
        except IntegrityError:
            messages.error(request, f"Cannot delete '{name}': it is used in one or more template mappings.")
        return redirect(reverse_lazy("field_templates:field_list"))


class FieldToggleActiveView(LoginRequiredMixin, StaffOrAdminMixin, View):
    def post(self, request, pk):
        field = get_object_or_404(StandardMasterField, pk=pk)
        field.is_active = not field.is_active
        field.save(update_fields=["is_active"])
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=request.user,
            action=ActivityLog.ACTION_UPDATE,
            description=f"Toggled Standard Field '{field.display_name}' active={field.is_active}",
            target=field,
            details={"is_active": field.is_active},
            request=request,
        )
        return redirect(reverse("field_templates:field_list"))


class FieldToggleDisplayedView(LoginRequiredMixin, StaffOrAdminMixin, View):
    def post(self, request, pk):
        field = get_object_or_404(StandardMasterField, pk=pk)
        field.is_displayed = not field.is_displayed
        field.save(update_fields=["is_displayed"])
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=request.user,
            action=ActivityLog.ACTION_UPDATE,
            description=f"Toggled Standard Field '{field.display_name}' displayed={field.is_displayed}",
            target=field,
            details={"is_displayed": field.is_displayed},
            request=request,
        )
        return redirect(reverse("field_templates:field_list"))


# ── Templates ─────────────────────────────────────────────────────────────────

class TemplateListView(LoginRequiredMixin, StaffOrAdminMixin, ListView):
    model = Template
    template_name = "field_templates/template_list.html"
    context_object_name = "templates"

    def get_queryset(self):
        return Template.objects.select_related("distributor").order_by("name")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        ctx["global_templates"] = qs.filter(scope="global")
        ctx["assigned_templates"] = qs.filter(scope="assigned")
        return ctx


class TemplateCreateView(LoginRequiredMixin, StaffOrAdminMixin, CreateView):
    model = Template
    form_class = TemplateForm
    template_name = "field_templates/template_form.html"

    def get_initial(self):
        initial = super().get_initial()
        distributor_pk = self.request.GET.get("distributor")
        if distributor_pk:
            initial["scope"] = "assigned"
            initial["distributor"] = distributor_pk
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        self.object = form.save()
        # Auto-create first version
        version = TemplateVersion.objects.create(
            template=self.object,
            version_number=1,
            is_active=True,
            created_by=self.request.user,
        )
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=self.request.user,
            action=ActivityLog.ACTION_CREATE,
            description=f"Created Template '{self.object.name}' (scope={self.object.scope})",
            target=self.object,
            details={"scope": self.object.scope, "distributor_id": self.object.distributor_id},
            request=self.request,
        )
        messages.success(self.request, "Template created. Now add field mappings.")
        return redirect(reverse("field_templates:version_mappings", kwargs={"pk": version.pk}))


class TemplateDetailView(LoginRequiredMixin, StaffOrAdminMixin, DetailView):
    model = Template
    template_name = "field_templates/template_detail.html"
    context_object_name = "template"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["versions"] = self.object.versions.order_by("-version_number")
        return ctx


class TemplateEditView(LoginRequiredMixin, StaffOrAdminMixin, UpdateView):
    model = Template
    form_class = TemplateForm
    template_name = "field_templates/template_form.html"

    def get_success_url(self):
        return reverse("field_templates:template_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=self.request.user,
            action=ActivityLog.ACTION_UPDATE,
            description=f"Updated Template '{self.object.name}'",
            target=self.object,
            details={"changed_fields": list(form.changed_data)},
            request=self.request,
        )
        messages.success(self.request, "Template updated.")
        return response


class TemplateDeleteView(LoginRequiredMixin, StaffOrAdminMixin, View):
    def post(self, request, pk):
        template = get_object_or_404(Template, pk=pk)
        name = template.name
        pk_val = template.pk
        # Check if any version is in use
        if template.versions.filter(processing_runs__isnull=False).exists():
            messages.error(request, f"Cannot delete '{name}': one or more versions are used in processing runs.")
            return redirect(reverse("field_templates:template_detail", kwargs={"pk": pk}))
        template.delete()
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=request.user,
            action=ActivityLog.ACTION_DELETE,
            description=f"Deleted Template '{name}'",
            details={"template_id": pk_val, "name": name},
            request=request,
        )
        messages.success(request, f"Template '{name}' deleted.")
        return redirect(reverse_lazy("field_templates:template_list"))


class VersionDeleteView(LoginRequiredMixin, StaffOrAdminMixin, View):
    def post(self, request, pk):
        version = get_object_or_404(TemplateVersion, pk=pk)
        template_pk = version.template_id
        template_name = version.template.name
        version_num = version.version_number
        if version.is_in_use:
            messages.error(request, f"Cannot delete v{version.version_number}: it is used in processing runs.")
            return redirect(reverse("field_templates:version_detail", kwargs={"pk": pk}))
        version.delete()
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=request.user,
            action=ActivityLog.ACTION_DELETE,
            description=f"Deleted Template '{template_name}' version {version_num}",
            details={"template_id": template_pk, "version": version_num},
            request=request,
        )
        messages.success(request, f"Version {version_num} deleted.")
        return redirect(reverse("field_templates:template_detail", kwargs={"pk": template_pk}))


class TemplateNewVersionView(LoginRequiredMixin, StaffOrAdminMixin, View):
    """Create a new TemplateVersion (cloning mappings from the latest)."""

    def post(self, request, pk):
        template = get_object_or_404(Template, pk=pk)
        latest = template.versions.order_by("-version_number").first()
        new_number = (latest.version_number + 1) if latest else 1
        new_version = TemplateVersion.objects.create(
            template=template,
            version_number=new_number,
            is_active=True,
            created_by=request.user,
            notes=request.POST.get("notes", ""),
        )
        # Clone mappings from latest version
        if latest:
            from .models import TemplateFieldMapping
            for mapping in latest.field_mappings.all():
                TemplateFieldMapping.objects.create(
                    template_version=new_version,
                    standard_field=mapping.standard_field,
                    source_column=mapping.source_column,
                )
        from apps.core.services import log_activity
        from apps.core.models import ActivityLog
        log_activity(
            user=request.user,
            action=ActivityLog.ACTION_CREATE,
            description=f"Created new version v{new_number} for Template '{template.name}'",
            target=new_version,
            details={"template_id": template.pk, "version": new_number},
            request=request,
        )
        messages.success(request, f"Version {new_number} created.")
        return redirect(reverse("field_templates:version_mappings", kwargs={"pk": new_version.pk}))


class VersionDetailView(LoginRequiredMixin, StaffOrAdminMixin, DetailView):
    model = TemplateVersion
    template_name = "field_templates/template_version_detail.html"
    context_object_name = "version"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["mappings"] = self.object.field_mappings.select_related("standard_field")
        ctx["header_mappings"] = self.object.header_mappings.select_related("standard_field")
        return ctx


class VersionMappingsView(LoginRequiredMixin, StaffOrAdminMixin, View):
    """Manage Table Column Mappings + Header Field Mappings for a version."""

    template_name = "field_templates/version_mappings.html"

    def _get_version(self, pk):
        return get_object_or_404(TemplateVersion, pk=pk)

    def _ctx(self, version, formset, header_formset):
        return {
            "version": version,
            "formset": formset,
            "header_formset": header_formset,
            "standard_fields": StandardMasterField.objects.order_by("order", "name"),
        }

    def _auto_clone_if_in_use(self, version, request):
        """If version is in use, clone it to a new version and redirect."""
        if not version.is_in_use:
            return None  # No clone needed

        from .models import TemplateFieldMapping, HeaderFieldMapping
        new_number = version.template.versions.order_by("-version_number").first().version_number + 1
        new_version = TemplateVersion.objects.create(
            template=version.template,
            version_number=new_number,
            is_active=True,
            created_by=request.user,
            notes=f"Auto-cloned from v{version.version_number} for editing",
        )
        # Clone table mappings
        for m in version.field_mappings.all():
            TemplateFieldMapping.objects.create(
                template_version=new_version,
                standard_field=m.standard_field,
                source_column=m.source_column,
            )
        # Clone header mappings
        for hm in version.header_mappings.all():
            HeaderFieldMapping.objects.create(
                template_version=new_version,
                standard_field=hm.standard_field,
                label=hm.label,
            )
        messages.info(
            request,
            f"Version {version.version_number} is in use. "
            f"Created new version v{new_number} for editing.",
        )
        return new_version

    def get(self, request, pk):
        version = self._get_version(pk)
        new_version = self._auto_clone_if_in_use(version, request)
        if new_version:
            return redirect(reverse("field_templates:version_mappings", kwargs={"pk": new_version.pk}))
        formset = TemplateFieldMappingFormSet(instance=version)
        header_formset = HeaderFieldMappingFormSet(instance=version, prefix="header")
        return render(request, self.template_name, self._ctx(version, formset, header_formset))

    def post(self, request, pk):
        version = self._get_version(pk)
        if version.is_in_use:
            messages.error(request, "Cannot save to a version that is in use. Edit the new version instead.")
            return redirect(reverse("field_templates:version_detail", kwargs={"pk": pk}))

        formset = TemplateFieldMappingFormSet(request.POST, instance=version)
        header_formset = HeaderFieldMappingFormSet(request.POST, instance=version, prefix="header")

        if formset.is_valid() and header_formset.is_valid():
            formset.save()
            header_formset.save()

            from apps.core.services import log_activity
            from apps.core.models import ActivityLog
            table_mappings = [
                {"field": m.standard_field.name, "source": m.source_column}
                for m in version.field_mappings.all()
            ]
            header_mappings = [
                {"field": m.standard_field.name, "label": m.label}
                for m in version.header_mappings.all()
            ]
            log_activity(
                user=request.user,
                action=ActivityLog.ACTION_UPDATE,
                description=f"Updated mappings for Template '{version.template.name}' v{version.version_number}",
                target=version,
                details={"table_mappings": table_mappings, "header_mappings": header_mappings},
                request=request,
            )
            messages.success(request, "Mappings saved.")
            return redirect(
                reverse("field_templates:template_detail", kwargs={"pk": version.template_id})
            )
        return render(request, self.template_name, self._ctx(version, formset, header_formset))


class FieldTemplateDownloadView(LoginRequiredMixin, StaffOrAdminMixin, View):
    """Download a sample .xlsx file with standard master field headers."""

    def get(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        fields = list(StandardMasterField.objects.order_by("order", "name"))
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Data Template"

        header_font = Font(bold=True, color="1B5E20")
        header_fill = PatternFill("solid", fgColor="E8F5E9")

        type_map = {
            "string": "text", "integer": "number", "decimal": "0.00",
            "date": "YYYY-MM-DD", "boolean": "true/false",
        }
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field.display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            ws.column_dimensions[cell.column_letter].width = max(len(field.display_name) + 4, 14)
            ws.cell(row=2, column=col_idx, value=type_map.get(field.data_type, field.data_type))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="master_data_template.xlsx"'
        return response
