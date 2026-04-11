"""Export service for Master Data records."""
from __future__ import annotations

import csv
import io
from datetime import datetime

from django.http import FileResponse, StreamingHttpResponse


def export_master_data(queryset, format_: str, standard_fields: list) -> FileResponse | StreamingHttpResponse:
    """Export a MasterDataRecord queryset to CSV or XLSX.

    Args:
        queryset: MasterDataRecord queryset (pre-filtered)
        format_: 'csv' or 'xlsx'
        standard_fields: list of StandardMasterField (ordered by .order)

    Returns:
        StreamingHttpResponse for CSV, FileResponse for XLSX.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"master_data_{timestamp}.{format_}"

    column_names = [sf.display_name for sf in standard_fields]

    if format_ == "csv":
        return _export_csv(queryset, column_names, standard_fields, filename)
    elif format_ == "xlsx":
        return _export_xlsx(queryset, column_names, standard_fields, filename)
    else:
        raise ValueError(f"Unsupported export format: {format_}")


def _row_to_values(record, standard_fields: list) -> list:
    return [record.data.get(sf.name, "") for sf in standard_fields]


class _EchoBuffer:
    """Minimal pseudo-buffer that returns written values (for StreamingHttpResponse)."""
    def write(self, value):
        return value


def _export_csv(queryset, column_names, standard_fields, filename) -> StreamingHttpResponse:
    pseudo_buffer = _EchoBuffer()
    writer = csv.writer(pseudo_buffer)

    def generate():
        yield writer.writerow(column_names)
        qs = queryset.select_related("distributor")
        for record in qs.iterator(chunk_size=500):
            yield writer.writerow(_row_to_values(record, standard_fields))

    response = StreamingHttpResponse(generate(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _export_xlsx(queryset, column_names, standard_fields, filename) -> FileResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Data"

    # Header row
    ws.append(column_names)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    qs = queryset.select_related("distributor")
    for record in qs.iterator(chunk_size=500):
        ws.append(_row_to_values(record, standard_fields))

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = FileResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
